"""
Менеджер для экспорта данных в Excel.
Реализует генерацию отчетов по оборудованию, материалам, монтажу и другим данным.
"""
import os
import tempfile
import asyncio
from typing import Dict, Any, List, Optional
from datetime import datetime, timedelta
from io import BytesIO
import structlog

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.context import AppContext
from storage.repositories.service_repository import ServiceRepository
from storage.repositories.installation_repository import InstallationRepository
from utils.date_utils import format_date
from modules.file.archive_manager import ArchiveManager


logger = structlog.get_logger(__name__)


class ExportManager:
    """Менеджер для экспорта данных в Excel."""
    
    def __init__(self, context: AppContext):
        self.context = context
        self.service_repository: Optional[ServiceRepository] = None
        self.installation_repository: Optional[InstallationRepository] = None
        self.archive_manager: Optional[ArchiveManager] = None
        self.temp_dir = tempfile.gettempdir()
    
    async def initialize(self) -> None:
        """Инициализирует менеджер экспорта."""
        self.service_repository = ServiceRepository(self.context.db_session)
        self.installation_repository = InstallationRepository(self.context.db_session)
        self.archive_manager = ArchiveManager(self.context)
        
        # Создаем директорию для временных файлов
        self.export_dir = os.path.join(self.temp_dir, 'ymk_exports')
        os.makedirs(self.export_dir, exist_ok=True)
        
        # Очищаем старые файлы при инициализации
        await self._cleanup_old_files()
        
        logger.info("ExportManager initialized")
    
    async def export_equipment(
        self,
        user_id: int,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Экспортирует данные об оборудовании в Excel.
        
        Args:
            user_id: ID пользователя
            filters: Фильтры для выборки оборудования
            
        Returns:
            Dict с информацией о созданном файле
        """
        try:
            # Проверяем права (только главный админ и админ)
            from modules.admin.admin_manager import AdminManager
            admin_manager = AdminManager(self.context)
            user_role = await admin_manager.get_user_role(user_id)
            
            if user_role not in ['main_admin', 'admin']:
                return {
                    'success': False,
                    'error': 'Недостаточно прав для экспорта'
                }
            
            # Получаем данные об оборудовании
            equipment_data = await self._get_equipment_data(filters)
            
            if not equipment_data:
                return {
                    'success': False,
                    'error': 'Нет данных для экспорта'
                }
            
            # Создаем Excel файл
            filename = f"equipment_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Оборудование"
            
            # Заголовки
            headers = [
                '№', 'Объект', 'Регион', 'Адрес', 'Наименование', 
                'Количество', 'Ед. изм.', 'Описание', 'Дата добавления'
            ]
            
            worksheet.append(headers)
            
            # Применяем стили к заголовкам
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Добавляем данные
            for idx, item in enumerate(equipment_data, 1):
                row = [
                    idx,
                    item.get('object_name', ''),
                    item.get('region_name', ''),
                    item.get('address', ''),
                    item.get('name', ''),
                    item.get('quantity', 0),
                    item.get('unit', 'шт.'),
                    item.get('description', ''),
                    format_date(item.get('created_at')) if item.get('created_at') else ''
                ]
                worksheet.append(row)
            
            # Настраиваем ширину колонок
            column_widths = [5, 30, 20, 40, 40, 10, 10, 50, 15]
            for i, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = width
            
            # Добавляем итоги
            total_row = len(equipment_data) + 3
            worksheet.cell(row=total_row, column=5, value="ИТОГО оборудование:")
            worksheet.cell(row=total_row, column=5).font = Font(bold=True)
            
            total_equipment = sum(item.get('quantity', 0) for item in equipment_data)
            worksheet.cell(row=total_row, column=6, value=total_equipment)
            worksheet.cell(row=total_row, column=6).font = Font(bold=True)
            
            # Сохраняем файл
            workbook.save(filepath)
            
            # Отправляем файл пользователю
            file_info = await self._send_excel_file(user_id, filepath, "Оборудование")
            
            # Сохраняем в архив
            await self._archive_export(filepath, 'equipment', user_id)
            
            return {
                'success': True,
                'file_info': file_info,
                'record_count': len(equipment_data),
                'total_equipment': total_equipment
            }
            
        except Exception as e:
            logger.error("Failed to export equipment", error=str(e))
            return {
                'success': False,
                'error': str(e)
            }
    
    async def export_materials(
        self,
        user_id: int,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Экспортирует данные о материалах в Excel.
        
        Args:
            user_id: ID пользователя
            filters: Фильтры для выборки материалов
            
        Returns:
            Dict с информацией о созданном файле
        """
        try:
            # Проверяем права
            from modules.admin.admin_manager import AdminManager
            admin_manager = AdminManager(self.context)
            user_role = await admin_manager.get_user_role(user_id)
            
            if user_role not in ['main_admin', 'admin']:
                return {
                    'success': False,
                    'error': 'Недостаточно прав для экспорта'
                }
            
            # Получаем данные о материалах
            materials_data = await self._get_materials_data(filters)
            
            if not materials_data:
                return {
                    'success': False,
                    'error': 'Нет данных для экспорта'
                }
            
            # Создаем Excel файл
            filename = f"materials_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            workbook = Workbook()
            
            # Лист общих материалов
            general_sheet = workbook.active
            general_sheet.title = "Общие материалы"
            
            # Заголовки для общего листа
            general_headers = [
                '№', 'Объект монтажа', 'Раздел', 'Наименование материала',
                'Количество', 'Ед. изм.', 'Описание', 'Плановый расход',
                'Фактический расход', 'Остаток', 'Дата добавления'
            ]
            
            general_sheet.append(general_headers)
            
            # Стили для заголовков
            self._apply_header_styles(general_sheet, len(general_headers))
            
            # Группируем материалы по разделам
            materials_by_section = {}
            for material in materials_data:
                section_name = material.get('section_name', 'Общее')
                if section_name not in materials_by_section:
                    materials_by_section[section_name] = []
                materials_by_section[section_name].append(material)
            
            # Добавляем общие материалы
            row_num = 2
            for material in materials_by_section.get('Общее', []):
                general_sheet.append([
                    row_num - 1,
                    material.get('object_name', ''),
                    material.get('section_name', ''),
                    material.get('name', ''),
                    material.get('quantity', 0),
                    material.get('unit', 'шт.'),
                    material.get('description', ''),
                    material.get('planned', 0),
                    material.get('actual', 0),
                    material.get('balance', 0),
                    format_date(material.get('created_at')) if material.get('created_at') else ''
                ])
                row_num += 1
            
            # Создаем листы для каждого раздела
            for section_name, section_materials in materials_by_section.items():
                if section_name == 'Общее':
                    continue
                
                sheet = workbook.create_sheet(title=section_name[:31])  # Ограничение длины названия листа
                
                section_headers = [
                    '№', 'Наименование материала', 'Количество', 'Ед. изм.',
                    'Описание', 'Плановый расход', 'Фактический расход', 'Остаток'
                ]
                
                sheet.append(section_headers)
                self._apply_header_styles(sheet, len(section_headers))
                
                for idx, material in enumerate(section_materials, 1):
                    sheet.append([
                        idx,
                        material.get('name', ''),
                        material.get('quantity', 0),
                        material.get('unit', 'шт.'),
                        material.get('description', ''),
                        material.get('planned', 0),
                        material.get('actual', 0),
                        material.get('balance', 0)
                    ])
            
            # Настраиваем ширину колонок для общего листа
            column_widths = [5, 30, 20, 40, 10, 10, 50, 15, 15, 15, 15]
            for i, width in enumerate(column_widths, 1):
                general_sheet.column_dimensions[get_column_letter(i)].width = width
            
            # Добавляем сводный лист
            summary_sheet = workbook.create_sheet(title="Сводка")
            
            summary_headers = ['Раздел', 'Количество материалов', 'Общее количество', 'Ед. изм.']
            summary_sheet.append(summary_headers)
            self._apply_header_styles(summary_sheet, len(summary_headers))
            
            for section_name, section_materials in materials_by_section.items():
                total_quantity = sum(m.get('quantity', 0) for m in section_materials)
                summary_sheet.append([
                    section_name,
                    len(section_materials),
                    total_quantity,
                    'шт.'  # Предполагаем одинаковые единицы измерения
                ])
            
            # Сохраняем файл
            workbook.save(filepath)
            
            # Отправляем файл пользователю
            file_info = await self._send_excel_file(user_id, filepath, "Материалы")
            
            # Сохраняем в архив
            await self._archive_export(filepath, 'materials', user_id)
            
            return {
                'success': True,
                'file_info': file_info,
                'record_count': len(materials_data),
                'sections_count': len(materials_by_section)
            }
            
        except Exception as e:
            logger.error("Failed to export materials", error=str(e))
            return {
                'success': False,
                'error': str(e)
            }
    
    async def export_installation(
        self,
        user_id: int,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Экспортирует данные о монтаже в Excel.
        
        Args:
            user_id: ID пользователя
            filters: Фильтры для выборки данных монтажа
            
        Returns:
            Dict с информацией о созданном файле
        """
        try:
            # Проверяем права
            from modules.admin.admin_manager import AdminManager
            admin_manager = AdminManager(self.context)
            user_role = await admin_manager.get_user_role(user_id)
            
            if user_role not in ['main_admin', 'admin']:
                return {
                    'success': False,
                    'error': 'Недостаточно прав для экспорта'
                }
            
            # Получаем данные о монтаже
            installation_data = await self._get_installation_data(filters)
            
            if not installation_data:
                return {
                    'success': False,
                    'error': 'Нет данных для экспорта'
                }
            
            # Создаем Excel файл
            filename = f"installation_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            workbook = Workbook()
            
            # Лист с основными данными объектов монтажа
            objects_sheet = workbook.active
            objects_sheet.title = "Объекты монтажа"
            
            objects_headers = [
                '№', 'Сокращенное название', 'Полное название', 'Адрес',
                'Контракт', 'Номер контракта', 'Дата контракта',
                'Дата начала', 'Дата окончания', 'Системы', 'Примечания',
                'Ответственный', 'Статус', 'Дата создания'
            ]
            
            objects_sheet.append(objects_headers)
            self._apply_header_styles(objects_sheet, len(objects_headers))
            
            for idx, obj in enumerate(installation_data.get('objects', []), 1):
                objects_sheet.append([
                    idx,
                    obj.get('short_name', ''),
                    obj.get('full_name', ''),
                    obj.get('address', ''),
                    obj.get('contract_type', ''),
                    obj.get('contract_number', ''),
                    format_date(obj.get('contract_date')) if obj.get('contract_date') else '',
                    format_date(obj.get('start_date')) if obj.get('start_date') else '',
                    format_date(obj.get('end_date')) if obj.get('end_date') else '',
                    obj.get('systems', ''),
                    obj.get('notes', ''),
                    obj.get('responsible', ''),
                    obj.get('status', ''),
                    format_date(obj.get('created_at')) if obj.get('created_at') else ''
                ])
            
            # Лист с проектами
            if installation_data.get('projects'):
                projects_sheet = workbook.create_sheet(title="Проекты")
                
                projects_headers = ['Объект', 'Название проекта', 'Описание', 'Файл', 'Дата добавления']
                projects_sheet.append(projects_headers)
                self._apply_header_styles(projects_sheet, len(projects_headers))
                
                for idx, project in enumerate(installation_data['projects'], 1):
                    projects_sheet.append([
                        project.get('object_name', ''),
                        project.get('name', ''),
                        project.get('description', ''),
                        'Есть' if project.get('has_file') else 'Нет',
                        format_date(project.get('created_at')) if project.get('created_at') else ''
                    ])
            
            # Лист с поставками
            if installation_data.get('supplies'):
                supplies_sheet = workbook.create_sheet(title="Поставки")
                
                supplies_headers = [
                    'Объект', 'Служба доставки', 'Дата доставки',
                    'Документ', 'Описание', 'Статус', 'Напоминание'
                ]
                
                supplies_sheet.append(supplies_headers)
                self._apply_header_styles(supplies_sheet, len(supplies_headers))
                
                for idx, supply in enumerate(installation_data['supplies'], 1):
                    supplies_sheet.append([
                        supply.get('object_name', ''),
                        supply.get('service', ''),
                        format_date(supply.get('delivery_date')) if supply.get('delivery_date') else '',
                        supply.get('document', ''),
                        supply.get('description', ''),
                        supply.get('status', ''),
                        'Есть' if supply.get('has_reminder') else 'Нет'
                    ])
            
            # Настраиваем ширину колонок
            column_widths = [5, 20, 40, 40, 15, 15, 15, 15, 15, 30, 50, 20, 15, 15]
            for i, width in enumerate(column_widths, 1):
                objects_sheet.column_dimensions[get_column_letter(i)].width = width
            
            # Сохраняем файл
            workbook.save(filepath)
            
            # Отправляем файл пользователю
            file_info = await self._send_excel_file(user_id, filepath, "Монтаж")
            
            # Сохраняем в архив
            await self._archive_export(filepath, 'installation', user_id)
            
            return {
                'success': True,
                'file_info': file_info,
                'objects_count': len(installation_data.get('objects', [])),
                'projects_count': len(installation_data.get('projects', [])),
                'supplies_count': len(installation_data.get('supplies', []))
            }
            
        except Exception as e:
            logger.error("Failed to export installation data", error=str(e))
            return {
                'success': False,
                'error': str(e)
            }
    
    async def export_all_data(
        self,
        user_id: int,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """
        Экспортирует все данные системы в Excel с разными листами.
        
        Args:
            user_id: ID пользователя
            start_date: Начальная дата периода
            end_date: Конечная дата периода
            
        Returns:
            Dict с информацией о созданном файле
        """
        try:
            # Проверяем права (только главный админ)
            from modules.admin.admin_manager import AdminManager
            admin_manager = AdminManager(self.context)
            user_role = await admin_manager.get_user_role(user_id)
            
            if user_role != 'main_admin':
                return {
                    'success': False,
                    'error': 'Только главный администратор может экспортировать все данные'
                }
            
            # Создаем Excel файл
            filename = f"full_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            workbook = Workbook()
            
            # Экспорт различных данных в разные листы
            export_tasks = [
                self._export_service_data(workbook, start_date, end_date),
                self._export_installation_data_sheets(workbook, start_date, end_date),
                self._export_problems_data(workbook, start_date, end_date),
                self._export_reminders_data(workbook, start_date, end_date),
                self._export_users_data(workbook),
                self._export_logs_data(workbook, start_date, end_date)
            ]
            
            # Выполняем все задачи экспорта
            results = await asyncio.gather(*export_tasks, return_exceptions=True)
            
            # Удаляем пустой лист по умолчанию
            if 'Sheet' in workbook.sheetnames:
                std = workbook['Sheet']
                workbook.remove(std)
            
            # Сохраняем файл
            workbook.save(filepath)
            
            # Отправляем файл пользователю
            file_info = await self._send_excel_file(user_id, filepath, "Полный экспорт")
            
            # Сохраняем в архив
            await self._archive_export(filepath, 'full_export', user_id)
            
            return {
                'success': True,
                'file_info': file_info,
                'sheets_count': len(workbook.sheetnames),
                'filename': filename
            }
            
        except Exception as e:
            logger.error("Failed to export all data", error=str(e))
            return {
                'success': False,
                'error': str(e)
            }
    
    async def _get_equipment_data(self, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """Получает данные об оборудовании."""
        if not self.service_repository:
            return []
        
        try:
            return await self.service_repository.get_equipment_for_export(filters)
        except Exception as e:
            logger.error("Failed to get equipment data", error=str(e))
            return []
    
    async def _get_materials_data(self, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """Получает данные о материалах."""
        if not self.installation_repository:
            return []
        
        try:
            return await self.installation_repository.get_materials_for_export(filters)
        except Exception as e:
            logger.error("Failed to get materials data", error=str(e))
            return []
    
    async def _get_installation_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Получает данные о монтаже."""
        if not self.installation_repository:
            return {'objects': [], 'projects': [], 'supplies': []}
        
        try:
            objects = await self.installation_repository.get_objects_for_export(filters)
            projects = await self.installation_repository.get_projects_for_export(filters)
            supplies = await self.installation_repository.get_supplies_for_export(filters)
            
            return {
                'objects': objects,
                'projects': projects,
                'supplies': supplies
            }
        except Exception as e:
            logger.error("Failed to get installation data", error=str(e))
            return {'objects': [], 'projects': [], 'supplies': []}
    
    async def _export_service_data(
        self,
        workbook: Workbook,
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> None:
        """Экспортирует данные обслуживания."""
        try:
            sheet = workbook.create_sheet(title="Обслуживание")
            
            headers = [
                'Регион', 'Объект', 'Адрес', 'Контракт', 'Номер контракта',
                'Дата контракта', 'Дата начала', 'Дата окончания',
                'Системы', 'ЗИП', 'Диспетчеризация', 'Примечания',
                'Ответственный', 'Статус'
            ]
            
            sheet.append(headers)
            self._apply_header_styles(sheet, len(headers))
            
            # Получаем данные
            if self.service_repository:
                service_data = await self.service_repository.get_service_for_export(start_date, end_date)
                
                for idx, item in enumerate(service_data, 1):
                    sheet.append([
                        item.get('region_name', ''),
                        item.get('object_name', ''),
                        item.get('address', ''),
                        item.get('contract_type', ''),
                        item.get('contract_number', ''),
                        format_date(item.get('contract_date')) if item.get('contract_date') else '',
                        format_date(item.get('start_date')) if item.get('start_date') else '',
                        format_date(item.get('end_date')) if item.get('end_date') else '',
                        item.get('systems', ''),
                        item.get('zip_payment', ''),
                        item.get('dispatching', ''),
                        item.get('notes', ''),
                        item.get('responsible', ''),
                        item.get('status', '')
                    ])
            
        except Exception as e:
            logger.error("Failed to export service data", error=str(e))
    
    async def _export_installation_data_sheets(
        self,
        workbook: Workbook,
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> None:
        """Экспортирует данные монтажа в несколько листов."""
        try:
            installation_data = await self._get_installation_data({
                'start_date': start_date,
                'end_date': end_date
            })
            
            # Лист объектов монтажа
            objects_sheet = workbook.create_sheet(title="Монтаж_Объекты")
            objects_headers = ['Объект', 'Адрес', 'Контракт', 'Статус', 'Ответственный']
            objects_sheet.append(objects_headers)
            self._apply_header_styles(objects_sheet, len(objects_headers))
            
            for obj in installation_data.get('objects', []):
                objects_sheet.append([
                    obj.get('short_name', ''),
                    obj.get('address', ''),
                    obj.get('contract_type', ''),
                    obj.get('status', ''),
                    obj.get('responsible', '')
                ])
            
            # Лист проектов
            if installation_data.get('projects'):
                projects_sheet = workbook.create_sheet(title="Монтаж_Проекты")
                projects_headers = ['Объект', 'Проект', 'Файл']
                projects_sheet.append(projects_headers)
                self._apply_header_styles(projects_sheet, len(projects_headers))
                
                for project in installation_data['projects']:
                    projects_sheet.append([
                        project.get('object_name', ''),
                        project.get('name', ''),
                        'Есть' if project.get('has_file') else 'Нет'
                    ])
            
        except Exception as e:
            logger.error("Failed to export installation data sheets", error=str(e))
    
    async def _export_problems_data(
        self,
        workbook: Workbook,
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> None:
        """Экспортирует данные о проблемах."""
        try:
            sheet = workbook.create_sheet(title="Проблемы")
            
            headers = [
                'Объект', 'Тип', 'Описание проблемы', 'Статус',
                'Дата создания', 'Дата решения', 'Решил', 'Решение'
            ]
            
            sheet.append(headers)
            self._apply_header_styles(sheet, len(headers))
            
            if self.service_repository:
                problems = await self.service_repository.get_problems_for_export(start_date, end_date)
                
                for problem in problems:
                    sheet.append([
                        problem.get('object_name', ''),
                        problem.get('type', ''),
                        problem.get('description', ''),
                        problem.get('status', ''),
                        format_date(problem.get('created_at')) if problem.get('created_at') else '',
                        format_date(problem.get('resolved_at')) if problem.get('resolved_at') else '',
                        problem.get('resolved_by', ''),
                        problem.get('solution', '')
                    ])
            
        except Exception as e:
            logger.error("Failed to export problems data", error=str(e))
    
    async def _export_reminders_data(
        self,
        workbook: Workbook,
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> None:
        """Экспортирует данные о напоминаниях."""
        try:
            sheet = workbook.create_sheet(title="Напоминания")
            
            headers = [
                'Объект', 'Тип объекта', 'Дата напоминания',
                'Текст напоминания', 'Статус', 'Автор'
            ]
            
            sheet.append(headers)
            self._apply_header_styles(sheet, len(headers))
            
            # Получаем данные из репозитория напоминаний
            from storage.repositories.reminder_repository import ReminderRepository
            reminder_repo = ReminderRepository(self.context.db_session)
            
            reminders = await reminder_repo.get_reminders_for_export(start_date, end_date)
            
            for reminder in reminders:
                sheet.append([
                    reminder.get('object_name', ''),
                    reminder.get('object_type', ''),
                    format_date(reminder.get('reminder_date')) if reminder.get('reminder_date') else '',
                    reminder.get('reminder_text', ''),
                    reminder.get('status', ''),
                    reminder.get('author_name', '')
                ])
            
        except Exception as e:
            logger.error("Failed to export reminders data", error=str(e))
    
    async def _export_users_data(self, workbook: Workbook) -> None:
        """Экспортирует данные о пользователях."""
        try:
            sheet = workbook.create_sheet(title="Пользователи")
            
            headers = [
                'ID', 'Имя', 'Username', 'Роль', 'Дата регистрации',
                'Последняя активность', 'Количество объектов'
            ]
            
            sheet.append(headers)
            self._apply_header_styles(sheet, len(headers))
            
            from storage.repositories.user_repository import UserRepository
            user_repo = UserRepository(self.context.db_session)
            
            users = await user_repo.get_users_for_export()
            
            for user in users:
                sheet.append([
                    user.get('user_id', ''),
                    user.get('full_name', ''),
                    user.get('username', ''),
                    user.get('role', ''),
                    format_date(user.get('created_at')) if user.get('created_at') else '',
                    format_date(user.get('last_active')) if user.get('last_active') else '',
                    user.get('objects_count', 0)
                ])
            
        except Exception as e:
            logger.error("Failed to export users data", error=str(e))
    
    async def _export_logs_data(
        self,
        workbook: Workbook,
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> None:
        """Экспортирует данные логов."""
        try:
            sheet = workbook.create_sheet(title="Логи")
            
            headers = [
                'Дата', 'Пользователь', 'Тип сущности', 'Сущность',
                'Действие', 'Изменения'
            ]
            
            sheet.append(headers)
            self._apply_header_styles(sheet, len(headers))
            
            from storage.repositories.log_repository import LogRepository
            log_repo = LogRepository(self.context.db_session)
            
            logs = await log_repo.get_logs_for_export(start_date, end_date)
            
            for log in logs:
                sheet.append([
                    format_date(log.get('timestamp')) if log.get('timestamp') else '',
                    log.get('user_name', ''),
                    log.get('entity_type', ''),
                    log.get('entity_name', ''),
                    log.get('action', ''),
                    log.get('changes', '')[:100]  # Ограничиваем длину
                ])
            
        except Exception as e:
            logger.error("Failed to export logs data", error=str(e))
    
    def _apply_header_styles(self, worksheet, header_count: int) -> None:
        """Применяет стили к заголовкам таблицы."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, header_count + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    async def _send_excel_file(
        self,
        user_id: int,
        filepath: str,
        export_type: str
    ) -> Dict[str, Any]:
        """
        Отправляет Excel файл пользователю.
        
        Args:
            user_id: ID пользователя
            filepath: Путь к файлу
            export_type: Тип экспорта
            
        Returns:
            Dict с информацией о файле
        """
        try:
            with open(filepath, 'rb') as file:
                # Отправляем файл пользователю
                message = await self.context.bot.send_document(
                    chat_id=user_id,
                    document=types.FSInputFile(filepath),
                    caption=f"📊 Экспорт данных: {export_type}\n"
                           f"📅 Дата: {format_date(datetime.now())}\n"
                           f"📄 Файл: {os.path.basename(filepath)}"
                )
            
            file_info = {
                'file_id': message.document.file_id,
                'file_name': os.path.basename(filepath),
                'file_size': os.path.getsize(filepath),
                'message_id': message.message_id
            }
            
            return file_info
            
        except Exception as e:
            logger.error("Failed to send Excel file", error=str(e))
            raise
    
    async def _archive_export(
        self,
        filepath: str,
        export_type: str,
        user_id: int
    ) -> None:
        """Сохраняет экспортированный файл в архив."""
        try:
            if not self.archive_manager:
                return
            
            # Получаем информацию о пользователе
            from modules.admin.admin_manager import AdminManager
            admin_manager = AdminManager(self.context)
            user_info = await admin_manager.get_user_info(user_id)
            
            # Читаем файл
            with open(filepath, 'rb') as file:
                file_data = file.read()
            
            # Сохраняем в архив
            await self.archive_manager.save_export_to_archive(
                file_data=file_data,
                file_name=os.path.basename(filepath),
                export_type=export_type,
                user_id=user_id,
                user_name=user_info.get('full_name', 'Неизвестно')
            )
            
            logger.info("Export archived", export_type=export_type, user_id=user_id)
            
        except Exception as e:
            logger.error("Failed to archive export", error=str(e))
    
    async def _cleanup_old_files(self, max_age_hours: int = 24) -> None:
        """Очищает старые временные файлы."""
        try:
            cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
            
            for filename in os.listdir(self.export_dir):
                filepath = os.path.join(self.export_dir, filename)
                
                if os.path.isfile(filepath):
                    file_mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
                    
                    if file_mtime < cutoff_time:
                        os.remove(filepath)
                        logger.debug("Old export file removed", filepath=filepath)
            
        except Exception as e:
            logger.error("Failed to cleanup old files", error=str(e))