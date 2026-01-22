"""
Сервис для экспорта данных в Excel.
Реализует экспорт оборудования, материалов, монтажа и других данных согласно ТЗ.
"""

import os
import logging
import tempfile
from typing import Dict, List, Optional, Any, Tuple, BinaryIO
from datetime import datetime, date
from enum import Enum

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils.date_utils import format_date

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Форматы экспорта."""
    EXCEL = "excel"
    CSV = "csv"
    PDF = "pdf"


class ExportType(str, Enum):
    """Типы экспорта."""
    EQUIPMENT = "equipment"
    MATERIALS = "materials"
    INSTALLATION = "installation"
    ALL_DATA = "all_data"
    SERVICE_OBJECTS = "service_objects"
    INSTALLATION_OBJECTS = "installation_objects"


class ExportService:
    """Сервис для экспорта данных в Excel."""
    
    # Стили для Excel
    HEADER_FONT = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    SUBHEADER_FONT = Font(name='Arial', size=11, bold=True, color='000000')
    SUBHEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    DATA_FONT = Font(name='Arial', size=10)
    DATA_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    def __init__(self, temp_dir: Optional[str] = None):
        self.temp_dir = temp_dir or tempfile.gettempdir()
    
    async def export_equipment(
        self,
        equipment_data: List[Dict[str, Any]],
        object_info: Optional[Dict[str, Any]] = None
    ) -> Tuple[str, str]:
        """
        Экспорт данных об оборудовании в Excel.
        
        Args:
            equipment_data: Список данных об оборудовании
            object_info: Информация об объекте
            
        Returns:
            Кортеж (путь к файлу, имя файла)
        """
        try:
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Оборудование"
            
            # Добавляем заголовок
            if object_info:
                title = f"Оборудование объекта: {object_info.get('short_name', 'Неизвестный объект')}"
            else:
                title = "Оборудование"
            
            ws.merge_cells('A1:H1')
            ws['A1'] = title
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Добавляем информацию об объекте
            if object_info:
                self._add_object_info(ws, object_info, start_row=3)
                data_start_row = 7
            else:
                data_start_row = 3
            
            # Заголовки таблицы
            headers = [
                "№",
                "Наименование",
                "Количество",
                "Единица измерения",
                "Адрес установки",
                "Описание",
                "Дата добавления",
                "Последнее изменение"
            ]
            
            # Записываем заголовки
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=data_start_row, column=col_idx, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.BORDER
            
            # Записываем данные
            for row_idx, equipment in enumerate(equipment_data, 1):
                row_num = data_start_row + row_idx
                
                ws.cell(row=row_num, column=1, value=row_idx).border = self.BORDER
                ws.cell(row=row_num, column=2, value=equipment.get('name', '')).border = self.BORDER
                ws.cell(row=row_num, column=3, value=equipment.get('quantity', 0)).border = self.BORDER
                ws.cell(row=row_num, column=4, value=equipment.get('unit', '')).border = self.BORDER
                ws.cell(row=row_num, column=5, value=equipment.get('address', '')).border = self.BORDER
                ws.cell(row=row_num, column=6, value=equipment.get('description', '')).border = self.BORDER
                
                # Форматируем даты
                created_at = equipment.get('created_at')
                if created_at:
                    if isinstance(created_at, datetime):
                        ws.cell(row=row_num, column=7, value=created_at.strftime('%d.%m.%Y')).border = self.BORDER
                    elif isinstance(created_at, str):
                        ws.cell(row=row_num, column=7, value=created_at).border = self.BORDER
                
                updated_at = equipment.get('updated_at')
                if updated_at:
                    if isinstance(updated_at, datetime):
                        ws.cell(row=row_num, column=8, value=updated_at.strftime('%d.%m.%Y')).border = self.BORDER
                    elif isinstance(updated_at, str):
                        ws.cell(row=row_num, column=8, value=updated_at).border = self.BORDER
            
            # Настраиваем ширину столбцов
            self._adjust_column_widths(ws)
            
            # Добавляем итоги
            total_row = data_start_row + len(equipment_data) + 1
            ws.cell(row=total_row, column=2, value="ИТОГО:").font = Font(bold=True)
            ws.cell(row=total_row, column=3, value=sum(e.get('quantity', 0) for e in equipment_data)).font = Font(bold=True)
            
            # Сохраняем файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"equipment_export_{timestamp}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"Equipment export created: {filename}")
            return filepath, filename
            
        except Exception as e:
            logger.error(f"Error exporting equipment: {e}")
            raise
    
    async def export_materials(
        self,
        materials_data: List[Dict[str, Any]],
        sections_data: Optional[List[Dict[str, Any]]] = None,
        object_info: Optional[Dict[str, Any]] = None
    ) -> Tuple[str, str]:
        """
        Экспорт данных о материалах монтажа.
        
        Args:
            materials_data: Список материалов
            sections_data: Данные о разделах (если есть)
            object_info: Информация об объекте
            
        Returns:
            Кортеж (путь к файлу, имя файла)
        """
        try:
            wb = Workbook()
            
            # Лист с общими материалами
            ws_general = wb.active
            ws_general.title = "Общие материалы"
            
            # Добавляем заголовок
            title = "Материалы монтажа"
            if object_info:
                title += f" - {object_info.get('short_name', 'Неизвестный объект')}"
            
            ws_general.merge_cells('A1:H1')
            ws_general['A1'] = title
            ws_general['A1'].font = Font(name='Arial', size=14, bold=True)
            ws_general['A1'].alignment = Alignment(horizontal='center')
            
            # Заголовки
            headers = [
                "№",
                "Наименование",
                "Общее количество",
                "Единица измерения",
                "Установлено",
                "Остаток",
                "Описание",
                "Раздел"
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_general.cell(row=3, column=col_idx, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.BORDER
            
            # Данные материалов
            for row_idx, material in enumerate(materials_data, 1):
                row_num = 3 + row_idx
                
                total_qty = material.get('quantity', 0)
                installed_qty = material.get('installed_quantity', 0)
                remaining = total_qty - installed_qty
                
                ws_general.cell(row=row_num, column=1, value=row_idx).border = self.BORDER
                ws_general.cell(row=row_num, column=2, value=material.get('name', '')).border = self.BORDER
                ws_general.cell(row=row_num, column=3, value=total_qty).border = self.BORDER
                ws_general.cell(row=row_num, column=4, value=material.get('unit', '')).border = self.BORDER
                ws_general.cell(row=row_num, column=5, value=installed_qty).border = self.BORDER
                ws_general.cell(row=row_num, column=6, value=remaining).border = self.BORDER
                ws_general.cell(row=row_num, column=7, value=material.get('description', '')).border = self.BORDER
                ws_general.cell(row=row_num, column=8, value=material.get('section_name', '')).border = self.BORDER
            
            # Итоги
            total_row = 3 + len(materials_data) + 1
            ws_general.cell(row=total_row, column=2, value="ИТОГО:").font = Font(bold=True)
            ws_general.cell(row=total_row, column=3, 
                          value=sum(m.get('quantity', 0) for m in materials_data)).font = Font(bold=True)
            ws_general.cell(row=total_row, column=5,
                          value=sum(m.get('installed_quantity', 0) for m in materials_data)).font = Font(bold=True)
            
            # Листы для разделов (если есть)
            if sections_data:
                for section in sections_data:
                    section_name = section.get('name', f"Раздел_{section.get('id', '')}")
                    # Ограничиваем имя листа 31 символом
                    sheet_name = str(section_name)[:31]
                    ws_section = wb.create_sheet(title=sheet_name)
                    
                    # Заголовок раздела
                    ws_section.merge_cells('A1:E1')
                    ws_section['A1'] = f"Раздел: {section_name}"
                    ws_section['A1'].font = Font(name='Arial', size=12, bold=True)
                    
                    # Материалы раздела
                    section_materials = section.get('materials', [])
                    if section_materials:
                        section_headers = ["№", "Наименование", "Количество", "Единица", "Установлено"]
                        
                        for col_idx, header in enumerate(section_headers, 1):
                            cell = ws_section.cell(row=3, column=col_idx, value=header)
                            cell.font = self.SUBHEADER_FONT
                            cell.fill = self.SUBHEADER_FILL
                            cell.border = self.BORDER
                        
                        for mat_idx, mat in enumerate(section_materials, 1):
                            row_num = 3 + mat_idx
                            ws_section.cell(row=row_num, column=1, value=mat_idx).border = self.BORDER
                            ws_section.cell(row=row_num, column=2, value=mat.get('name', '')).border = self.BORDER
                            ws_section.cell(row=row_num, column=3, value=mat.get('quantity', 0)).border = self.BORDER
                            ws_section.cell(row=row_num, column=4, value=mat.get('unit', '')).border = self.BORDER
                            ws_section.cell(row=row_num, column=5, value=mat.get('installed_quantity', 0)).border = self.BORDER
            
            # Настраиваем ширину столбцов
            for ws in wb.worksheets:
                self._adjust_column_widths(ws)
            
            # Сохраняем файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"materials_export_{timestamp}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"Materials export created: {filename}")
            return filepath, filename
            
        except Exception as e:
            logger.error(f"Error exporting materials: {e}")
            raise
    
    async def export_installation_data(
        self,
        objects_data: List[Dict[str, Any]],
        detailed: bool = False
    ) -> Tuple[str, str]:
        """
        Экспорт данных о монтаже.
        
        Args:
            objects_data: Список объектов монтажа
            detailed: Детальный экспорт со всеми данными
            
        Returns:
            Кортеж (путь к файлу, имя файла)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Объекты монтажа"
            
            # Заголовок
            ws.merge_cells('A1:L1')
            ws['A1'] = "Объекты монтажа - Экспорт данных"
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Информация о экспорте
            export_date = datetime.now().strftime("%d.%m.%Y %H:%M")
            ws['A2'] = f"Дата экспорта: {export_date}"
            ws['A3'] = f"Количество объектов: {len(objects_data)}"
            
            # Заголовки таблицы
            headers = [
                "№",
                "Сокращенное название",
                "Полное название",
                "Адрес",
                "Контракт №",
                "Дата контракта",
                "Сроки",
                "Системы",
                "Проектов",
                "Материалов",
                "Установлено %",
                "Статус"
            ]
            
            start_row = 5
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.BORDER
            
            # Данные объектов
            for obj_idx, obj in enumerate(objects_data, 1):
                row_num = start_row + obj_idx
                
                # Сроки контракта
                start_date = obj.get('start_date')
                end_date = obj.get('end_date')
                dates_str = ""
                if start_date and end_date:
                    if isinstance(start_date, datetime):
                        start_str = start_date.strftime('%d.%m.%Y')
                    else:
                        start_str = str(start_date)
                    
                    if isinstance(end_date, datetime):
                        end_str = end_date.strftime('%d.%m.%Y')
                    else:
                        end_str = str(end_date)
                    
                    dates_str = f"{start_str} - {end_str}"
                
                # Процент установки
                total_materials = obj.get('total_materials', 0)
                installed_materials = obj.get('installed_materials', 0)
                install_percent = 0
                if total_materials > 0:
                    install_percent = round((installed_materials / total_materials) * 100, 1)
                
                # Статус
                status = "Активный"
                if end_date:
                    if isinstance(end_date, datetime) and end_date < datetime.now():
                        status = "Завершен"
                    elif isinstance(end_date, str):
                        # Пытаемся парсить строку
                        try:
                            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                            if end_dt < datetime.now():
                                status = "Завершен"
                        except:
                            pass
                
                ws.cell(row=row_num, column=1, value=obj_idx).border = self.BORDER
                ws.cell(row=row_num, column=2, value=obj.get('short_name', '')).border = self.BORDER
                ws.cell(row=row_num, column=3, value=obj.get('full_name', '')).border = self.BORDER
                ws.cell(row=row_num, column=4, value=obj.get('address', '')).border = self.BORDER
                ws.cell(row=row_num, column=5, value=obj.get('contract_number', '')).border = self.BORDER
                
                # Форматируем дату контракта
                contract_date = obj.get('contract_date')
                if contract_date:
                    if isinstance(contract_date, datetime):
                        ws.cell(row=row_num, column=6, 
                              value=contract_date.strftime('%d.%m.%Y')).border = self.BORDER
                    else:
                        ws.cell(row=row_num, column=6, value=str(contract_date)).border = self.BORDER
                
                ws.cell(row=row_num, column=7, value=dates_str).border = self.BORDER
                
                # Системы
                systems = obj.get('systems', [])
                ws.cell(row=row_num, column=8, value=', '.join(systems)).border = self.BORDER
                
                ws.cell(row=row_num, column=9, value=obj.get('project_count', 0)).border = self.BORDER
                ws.cell(row=row_num, column=10, value=total_materials).border = self.BORDER
                ws.cell(row=row_num, column=11, value=f"{install_percent}%").border = self.BORDER
                ws.cell(row=row_num, column=12, value=status).border = self.BORDER
            
            # Детальный экспорт (дополнительные листы)
            if detailed:
                self._add_detailed_installation_sheets(wb, objects_data)
            
            # Настраиваем ширину столбцов
            self._adjust_column_widths(ws)
            
            # Итоги
            total_row = start_row + len(objects_data) + 1
            ws.cell(row=total_row, column=2, value="ИТОГО:").font = Font(bold=True)
            ws.cell(row=total_row, column=9, 
                   value=sum(o.get('project_count', 0) for o in objects_data)).font = Font(bold=True)
            ws.cell(row=total_row, column=10,
                   value=sum(o.get('total_materials', 0) for o in objects_data)).font = Font(bold=True)
            
            # Сохраняем файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"installation_export_{timestamp}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"Installation export created: {filename}")
            return filepath, filename
            
        except Exception as e:
            logger.error(f"Error exporting installation data: {e}")
            raise
    
    async def export_all_data(
        self,
        service_data: Dict[str, Any],
        installation_data: Dict[str, Any],
        user_data: Optional[Dict[str, Any]] = None
    ) -> Tuple[str, str]:
        """
        Комплексный экспорт всех данных.
        
        Args:
            service_data: Данные обслуживания
            installation_data: Данные монтажа
            user_data: Данные пользователей (опционально)
            
        Returns:
            Кортеж (путь к файлу, имя файла)
        """
        try:
            wb = Workbook()
            
            # 1. Сводка
            ws_summary = wb.active
            ws_summary.title = "Сводка"
            
            export_date = datetime.now().strftime("%d.%m.%Y %H:%M")
            ws_summary['A1'] = "КОМПЛЕКСНЫЙ ЭКСПОРТ ДАННЫХ"
            ws_summary['A1'].font = Font(name='Arial', size=16, bold=True)
            
            ws_summary['A3'] = f"Дата экспорта: {export_date}"
            ws_summary['A4'] = "СИСТЕМА УЧЕТА ЭЛЕКТРИКИ"
            
            # Статистика обслуживания
            ws_summary['A6'] = "ОБСЛУЖИВАНИЕ:"
            ws_summary['A6'].font = Font(bold=True)
            
            service_stats = service_data.get('statistics', {})
            ws_summary['A7'] = f"Регионов: {service_stats.get('regions_count', 0)}"
            ws_summary['A8'] = f"Объектов: {service_stats.get('objects_count', 0)}"
            ws_summary['A9'] = f"Проблем: {service_stats.get('problems_count', 0)}"
            ws_summary['A10'] = f"ТО: {service_stats.get('maintenance_count', 0)}"
            
            # Статистика монтажа
            ws_summary['A12'] = "МОНТАЖ:"
            ws_summary['A12'].font = Font(bold=True)
            
            installation_stats = installation_data.get('statistics', {})
            ws_summary['A13'] = f"Объектов: {installation_stats.get('objects_count', 0)}"
            ws_summary['A14'] = f"Проектов: {installation_stats.get('projects_count', 0)}"
            ws_summary['A15'] = f"Материалов: {installation_stats.get('materials_count', 0)}"
            
            # 2. Обслуживание
            if 'regions' in service_data:
                ws_service = wb.create_sheet("Регионы обслуживания")
                self._add_service_regions_sheet(ws_service, service_data.get('regions', []))
            
            if 'objects' in service_data:
                ws_service_objects = wb.create_sheet("Объекты обслуживания")
                self._add_service_objects_sheet(ws_service_objects, service_data.get('objects', []))
            
            # 3. Монтаж
            if 'objects' in installation_data:
                ws_installation = wb.create_sheet("Объекты монтажа")
                self._add_installation_objects_sheet(ws_installation, installation_data.get('objects', []))
            
            # 4. Пользователи (если есть)
            if user_data and 'users' in user_data:
                ws_users = wb.create_sheet("Пользователи")
                self._add_users_sheet(ws_users, user_data.get('users', []))
            
            # Удаляем дефолтный лист если он пустой
            if len(wb.worksheets) > 1 and ws_summary.max_row == 1:
                wb.remove(ws_summary)
            
            # Настраиваем ширину столбцов
            for ws in wb.worksheets:
                self._adjust_column_widths(ws)
            
            # Сохраняем файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"full_export_{timestamp}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"Full data export created: {filename}")
            return filepath, filename
            
        except Exception as e:
            logger.error(f"Error exporting all data: {e}")
            raise
    
    def _add_object_info(self, ws, object_info: Dict[str, Any], start_row: int = 1):
        """Добавление информации об объекте."""
        ws.cell(row=start_row, column=1, value="Информация об объекте:").font = Font(bold=True)
        
        info_rows = [
            ("Сокращенное название:", object_info.get('short_name', '')),
            ("Полное название:", object_info.get('full_name', '')),
            ("Контракт:", f"{object_info.get('contract_type', '')} №{object_info.get('contract_number', '')}"),
            ("Адрес:", object_info.get('address', '')),
        ]
        
        for i, (label, value) in enumerate(info_rows, 1):
            ws.cell(row=start_row + i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=start_row + i, column=2, value=value)
    
    def _adjust_column_widths(self, ws):
        """Автонастройка ширины столбцов."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_detailed_installation_sheets(self, wb: Workbook, objects_data: List[Dict[str, Any]]):
        """Добавление детальных листов для монтажа."""
        for obj in objects_data:
            obj_name = obj.get('short_name', f"Object_{obj.get('id', '')}")
            sheet_name = str(obj_name)[:31]
            
            try:
                ws_obj = wb.create_sheet(title=sheet_name)
                
                # Заголовок
                ws_obj.merge_cells('A1:E1')
                ws_obj['A1'] = f"Объект: {obj_name}"
                ws_obj['A1'].font = Font(size=12, bold=True)
                
                # Основная информация
                ws_obj['A3'] = "Основная информация:"
                ws_obj['A3'].font = Font(bold=True)
                
                info = [
                    ("Полное название:", obj.get('full_name', '')),
                    ("Адрес:", obj.get('address', '')),
                    ("Контракт:", obj.get('contract_number', '')),
                    ("Системы:", ', '.join(obj.get('systems', []))),
                ]
                
                for i, (label, value) in enumerate(info):
                    ws_obj.cell(row=4 + i, column=1, value=label).font = Font(bold=True)
                    ws_obj.cell(row=4 + i, column=2, value=value)
                
            except Exception as e:
                logger.warning(f"Could not create sheet for object {obj_name}: {e}")
    
    def _add_service_regions_sheet(self, ws, regions: List[Dict[str, Any]]):
        """Добавление листа с регионами обслуживания."""
        ws.title = "Регионы обслуживания"
        
        # Заголовок
        ws.merge_cells('A1:C1')
        ws['A1'] = "Регионы обслуживания"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Заголовки таблицы
        headers = ["№", "Сокращенное название", "Полное название", "Количество объектов"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Данные
        for idx, region in enumerate(regions, 1):
            row = 3 + idx
            ws.cell(row=row, column=1, value=idx).border = self.BORDER
            ws.cell(row=row, column=2, value=region.get('short_name', '')).border = self.BORDER
            ws.cell(row=row, column=3, value=region.get('full_name', '')).border = self.BORDER
            ws.cell(row=row, column=4, value=region.get('object_count', 0)).border = self.BORDER
    
    def _add_service_objects_sheet(self, ws, objects: List[Dict[str, Any]]):
        """Добавление листа с объектами обслуживания."""
        ws.title = "Объекты обслуживания"
        
        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = "Объекты обслуживания"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Заголовки
        headers = ["№", "Название", "Регион", "Контракт", "Адрес", "Системы"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Данные
        for idx, obj in enumerate(objects, 1):
            row = 3 + idx
            ws.cell(row=row, column=1, value=idx).border = self.BORDER
            ws.cell(row=row, column=2, value=obj.get('short_name', '')).border = self.BORDER
            ws.cell(row=row, column=3, value=obj.get('region_name', '')).border = self.BORDER
            ws.cell(row=row, column=4, value=obj.get('contract_number', '')).border = self.BORDER
            ws.cell(row=row, column=5, value=obj.get('address', '')).border = self.BORDER
            ws.cell(row=row, column=6, value=', '.join(obj.get('systems', []))).border = self.BORDER
    
    def _add_installation_objects_sheet(self, ws, objects: List[Dict[str, Any]]):
        """Добавление листа с объектами монтажа."""
        ws.title = "Объекты монтажа"
        
        # Заголовок
        ws.merge_cells('A1:G1')
        ws['A1'] = "Объекты монтажа"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Заголовки
        headers = ["№", "Название", "Контракт", "Адрес", "Проектов", "Материалов", "Статус"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Данные
        for idx, obj in enumerate(objects, 1):
            row = 3 + idx
            ws.cell(row=row, column=1, value=idx).border = self.BORDER
            ws.cell(row=row, column=2, value=obj.get('short_name', '')).border = self.BORDER
            ws.cell(row=row, column=3, value=obj.get('contract_number', '')).border = self.BORDER
            ws.cell(row=row, column=4, value=obj.get('address', '')).border = self.BORDER
            ws.cell(row=row, column=5, value=obj.get('project_count', 0)).border = self.BORDER
            ws.cell(row=row, column=6, value=obj.get('material_count', 0)).border = self.BORDER
            ws.cell(row=row, column=7, value=obj.get('status', 'Активный')).border = self.BORDER
    
    def _add_users_sheet(self, ws, users: List[Dict[str, Any]]):
        """Добавление листа с пользователями."""
        ws.title = "Пользователи"
        
        # Заголовок
        ws.merge_cells('A1:E1')
        ws['A1'] = "Пользователи системы"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Заголовки
        headers = ["№", "ID", "Username", "Имя", "Роль"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Данные
        for idx, user in enumerate(users, 1):
            row = 3 + idx
            ws.cell(row=row, column=1, value=idx).border = self.BORDER
            ws.cell(row=row, column=2, value=user.get('telegram_id', '')).border = self.BORDER
            ws.cell(row=row, column=3, value=user.get('username', '')).border = self.BORDER
            ws.cell(row=row, column=4, value=user.get('full_name', '')).border = self.BORDER
            ws.cell(row=row, column=5, value=user.get('role', 'Пользователь')).border = self.BORDER
    
    def cleanup_temp_files(self, older_than_hours: int = 24):
        """
        Очистка временных файлов экспорта.
        
        Args:
            older_than_hours: Удалять файлы старше указанного количества часов
        """
        try:
            import glob
            import time
            
            pattern = os.path.join(self.temp_dir, "*_export_*.xlsx")
            files = glob.glob(pattern)
            
            current_time = time.time()
            cutoff_time = current_time - (older_than_hours * 3600)
            
            deleted_count = 0
            for filepath in files:
                try:
                    file_time = os.path.getmtime(filepath)
                    if file_time < cutoff_time:
                        os.remove(filepath)
                        deleted_count += 1
                except Exception as e:
                    logger.warning(f"Could not delete file {filepath}: {e}")
            
            logger.info(f"Cleaned up {deleted_count} temporary export files")
            
        except Exception as e:
            logger.error(f"Error cleaning up temp files: {e}")